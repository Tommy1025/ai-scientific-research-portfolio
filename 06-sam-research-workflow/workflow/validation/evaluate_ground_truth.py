#!/usr/bin/env python3
"""Validate an AI-extracted SAM workbook against a reference workbook.

Rows are paired by DOI + normalized material identity. When one material appears
more than once in the same paper, the one-to-one pairing minimizes PCE
difference only to disambiguate repeated experimental rows. PCE is therefore not
reported as an independent accuracy metric.

The script reports exact field agreement on the paired records. For SMILES,
chemical identity is compared after RDKit canonicalization.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import unicodedata
from collections import defaultdict
from itertools import combinations, permutations
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from rdkit import Chem


PROCESS_FIELDS = [
    "NiO2",
    "ethanol",
    "toluene",
    "IPA",
    "THF",
    "chlorobenzene",
    "2-Methoxyethanol",
    "CH2CL2",
    "concentration(mg/ml)",
    "wash",
    "E",
    "Cs",
    "FA",
    "MA",
    "Pb",
    "Sn",
    "I",
    "Br",
    "CL",
    "C60",
    "BCP",
    "PC60BM",
    "PCBM",
    "PC61BM",
    "PEAI",
    "ALD-SnO2",
]

GROUPS = {
    "binary_process": ["NiO2", "wash"],
    "solvents": [
        "ethanol",
        "toluene",
        "IPA",
        "THF",
        "chlorobenzene",
        "2-Methoxyethanol",
        "CH2CL2",
    ],
    "A_site": ["Cs", "FA", "MA"],
    "B_site": ["Pb", "Sn"],
    "X_site": ["I", "Br", "CL"],
    "ETL_interface": [
        "C60",
        "BCP",
        "PC60BM",
        "PCBM",
        "PC61BM",
        "PEAI",
        "ALD-SnO2",
    ],
}

FIELDS = ["smile", *PROCESS_FIELDS]


def load_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["主表"] if "主表" in workbook.sheetnames else workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise ValueError(f"No rows found in {path}")
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    rows = [
        dict(zip(headers, row, strict=True))
        for row in values[1:]
        if any(value is not None for value in row)
    ]
    required = {"SAM/HTL材料名稱", "Reference_DOI", "PCE", *FIELDS}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"{path.name} is missing columns: {missing}")
    return headers, rows


def blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def normalize_text(value: object) -> str:
    if blank(value):
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    return re.sub(r"\s+", "", text)


def normalize_doi(value: object) -> str:
    text = normalize_text(value)
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if text.startswith(prefix):
            text = text[len(prefix) :]
    return text.strip()


def normalize_material(value: object) -> str:
    """Normalize dataset-specific process annotations around a material name."""
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"^niox\+", "", text)
    text = re.sub(r"[\s_\-+]+", "", text)
    return text


def numeric(value: object) -> float | None:
    if blank(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def equal_value(left: object, right: object, tolerance: float = 1e-9) -> bool:
    if blank(left) or blank(right):
        return blank(left) and blank(right)
    left_number = numeric(left)
    right_number = numeric(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) <= tolerance
    return normalize_text(left) == normalize_text(right)


def canonical_smiles(value: object) -> str | None:
    if blank(value):
        return None
    molecule = Chem.MolFromSmiles(str(value).strip())
    if molecule is None:
        return None
    return Chem.MolToSmiles(molecule, canonical=True, isomericSmiles=True)


def smiles_equal(left: object, right: object) -> bool:
    left_canonical = canonical_smiles(left)
    right_canonical = canonical_smiles(right)
    return (
        left_canonical is not None
        and right_canonical is not None
        and left_canonical == right_canonical
    )


def pce_distance(left: dict[str, object], right: dict[str, object]) -> float:
    left_value = numeric(left["PCE"])
    right_value = numeric(right["PCE"])
    if left_value is None or right_value is None:
        return 1_000_000.0
    return abs(left_value - right_value)


def pair_group(
    reference_rows: list[dict[str, object]],
    prediction_rows: list[dict[str, object]],
) -> list[tuple[dict[str, object], dict[str, object]]]:
    """Find the minimum-PCE-distance one-to-one pairing for one material group."""
    if len(reference_rows) <= len(prediction_rows):
        best: tuple[
            float, list[tuple[dict[str, object], dict[str, object]]]
        ] | None = None
        for selected in combinations(range(len(prediction_rows)), len(reference_rows)):
            for permuted in permutations(selected):
                pairs = [
                    (reference_rows[i], prediction_rows[permuted[i]])
                    for i in range(len(reference_rows))
                ]
                cost = sum(pce_distance(left, right) for left, right in pairs)
                if best is None or cost < best[0]:
                    best = (cost, pairs)
    else:
        best = None
        for selected in combinations(range(len(reference_rows)), len(prediction_rows)):
            for permuted in permutations(selected):
                pairs = [
                    (reference_rows[permuted[i]], prediction_rows[i])
                    for i in range(len(prediction_rows))
                ]
                cost = sum(pce_distance(left, right) for left, right in pairs)
                if best is None or cost < best[0]:
                    best = (cost, pairs)
    assert best is not None
    return best[1]


def build_pairs(
    reference_rows: Iterable[dict[str, object]],
    prediction_rows: Iterable[dict[str, object]],
) -> list[tuple[dict[str, object], dict[str, object]]]:
    reference_groups: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    prediction_groups: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)

    for row in reference_rows:
        key = (
            normalize_doi(row["Reference_DOI"]),
            normalize_material(row["SAM/HTL材料名稱"]),
        )
        reference_groups[key].append(row)
    for row in prediction_rows:
        key = (
            normalize_doi(row["Reference_DOI"]),
            normalize_material(row["SAM/HTL材料名稱"]),
        )
        prediction_groups[key].append(row)

    pairs: list[tuple[dict[str, object], dict[str, object]]] = []
    for key in sorted(reference_groups.keys() & prediction_groups.keys()):
        pairs.extend(pair_group(reference_groups[key], prediction_groups[key]))
    return pairs


def prf(
    correct: int,
    predicted: int,
    reference: int,
) -> tuple[float | None, float | None, float | None]:
    precision = correct / predicted if predicted else None
    recall = correct / reference if reference else None
    if precision is None or recall is None or precision + recall == 0:
        f1 = None
    else:
        f1 = 2 * precision * recall / (precision + recall)
    return precision, recall, f1


def percent(value: float | None) -> str:
    return "" if value is None else f"{100 * value:.3f}"


def evaluate_field(
    field: str,
    pairs: list[tuple[dict[str, object], dict[str, object]]],
) -> dict[str, object]:
    reference_nonblank = 0
    prediction_nonblank = 0
    correct = 0

    for reference_row, prediction_row in pairs:
        reference_value = reference_row[field]
        prediction_value = prediction_row[field]
        reference_has_value = not blank(reference_value)
        prediction_has_value = not blank(prediction_value)

        reference_nonblank += int(reference_has_value)
        prediction_nonblank += int(prediction_has_value)

        if reference_has_value and prediction_has_value:
            is_correct = (
                smiles_equal(reference_value, prediction_value)
                if field == "smile"
                else equal_value(reference_value, prediction_value)
            )
            correct += int(is_correct)

    precision, recall, f1 = prf(correct, prediction_nonblank, reference_nonblank)
    return {
        "type": "field",
        "name": field,
        "correct": correct,
        "prediction_nonblank": prediction_nonblank,
        "reference_nonblank": reference_nonblank,
        "precision_percent": percent(precision),
        "recall_percent": percent(recall),
        "f1_percent": percent(f1),
        "accuracy_percent": percent(
            correct / reference_nonblank if reference_nonblank else None
        ),
        "notes": (
            "SMILES compared by RDKit canonical identity."
            if field == "smile"
            else "Exact value agreement."
        ),
    }


def evaluate_group(
    name: str,
    fields: list[str],
    pairs: list[tuple[dict[str, object], dict[str, object]]],
) -> dict[str, object]:
    correct = 0
    total = len(fields) * len(pairs)
    for reference_row, prediction_row in pairs:
        for field in fields:
            correct += int(equal_value(reference_row[field], prediction_row[field]))
    return {
        "type": "group",
        "name": name,
        "correct": correct,
        "prediction_nonblank": "",
        "reference_nonblank": total,
        "precision_percent": "",
        "recall_percent": "",
        "f1_percent": "",
        "accuracy_percent": percent(correct / total if total else None),
        "notes": f"Micro exact-match accuracy across: {', '.join(fields)}",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ground-truth", required=True, type=Path)
    parser.add_argument("--prediction", required=True, type=Path)
    parser.add_argument("--out", type=Path, default=Path("validation_metrics.csv"))
    args = parser.parse_args()

    _, reference_rows = load_rows(args.ground_truth)
    _, prediction_rows = load_rows(args.prediction)
    pairs = build_pairs(reference_rows, prediction_rows)

    output_rows: list[dict[str, object]] = [
        {
            "type": "summary",
            "name": "paired_records",
            "correct": len(pairs),
            "prediction_nonblank": "",
            "reference_nonblank": "",
            "precision_percent": "",
            "recall_percent": "",
            "f1_percent": "",
            "accuracy_percent": "",
            "notes": (
                "Matched by DOI + normalized material identity; repeated materials "
                "paired by minimum PCE distance."
            ),
        }
    ]
    output_rows.extend(
        evaluate_group(name, fields, pairs) for name, fields in GROUPS.items()
    )
    output_rows.extend(evaluate_field(field, pairs) for field in FIELDS)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "type",
        "name",
        "correct",
        "prediction_nonblank",
        "reference_nonblank",
        "precision_percent",
        "recall_percent",
        "f1_percent",
        "accuracy_percent",
        "notes",
    ]
    with args.out.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(output_rows)

    print(f"paired_records={len(pairs)}")
    for row in output_rows:
        if row["type"] == "group":
            print(f"{row['name']}: {row['accuracy_percent']}%")
    smiles = next(
        row
        for row in output_rows
        if row["type"] == "field" and row["name"] == "smile"
    )
    print(
        "SMILES: "
        f"precision={smiles['precision_percent']}%, "
        f"recall={smiles['recall_percent']}%, "
        f"F1={smiles['f1_percent']}%"
    )
    print(f"wrote={args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
