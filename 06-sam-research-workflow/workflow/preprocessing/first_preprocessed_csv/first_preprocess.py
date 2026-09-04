#!/usr/bin/env python3
r"""Repeat the first XLSX-to-CSV preprocessing defined by FIRST_PREPROCESSING_GUIDE.md.

The script is intentionally conservative: it never modifies source workbooks,
requires explicit XLSX paths, stops on structural/data-contract errors, and
records every filtered or transformed row in a JSON audit report.

"""


from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import sys
import tempfile
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook


VERSION = "1.4.0"
TOLERANCE = Decimal("0.000001")
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
CSV_OUTPUT_DIR = SCRIPT_DIR / "csv"
JSON_OUTPUT_DIR = SCRIPT_DIR / "json"
PAPER_DIR_PATTERN = re.compile(r"^\d{4}_.+")

COLUMNS = [
    "編號",
    "SAM/HTL材料名稱",
    "smile",
    "NiO2",
    "ethanol",
    "toluene",
    "IPA",
    "THF",
    "chlorobenzene",
    "2-Methoxyethanol",
    "CH2CL2",
    "concentration(mg/ml)",
    "wash",
    "E",
    "Cs",
    "FA",
    "MA",
    "Pb",
    "Sn",
    "I",
    "Br",
    "CL",
    "C60",
    "BCP",
    "PC60BM",
    "PCBM",
    "PC61BM",
    "PEAI",
    "ALD-SnO2",
    "PCE",
    "Reference_DOI",
    "Ref_author",
    "Ref_journal",
    "Data_status",
    "Notes",
]

SOLVENT_COLUMNS = [
    "ethanol",
    "toluene",
    "IPA",
    "THF",
    "chlorobenzene",
    "2-Methoxyethanol",
    "CH2CL2",
]
X_SITE_COLUMNS = ["I", "Br", "CL"]
A_SITE_COLUMNS = ["Cs", "FA", "MA"]
B_SITE_COLUMNS = ["Pb", "Sn"]
ETL_COLUMNS = ["C60", "BCP", "PC60BM", "PCBM", "PC61BM", "PEAI", "ALD-SnO2"]
BINARY_COLUMNS = [
    "NiO2",
    "wash",
    "C60",
    "BCP",
    "PC60BM",
    "PCBM",
    "PC61BM",
    "PEAI",
    "ALD-SnO2",
]
REQUIRED_COLUMNS = [
    "smile",
    "NiO2",
    *SOLVENT_COLUMNS,
    "Cs",
    "FA",
    "MA",
    "Pb",
    "Sn",
    *X_SITE_COLUMNS,
    "C60",
    "BCP",
    "PC60BM",
    "PCBM",
    "PC61BM",
    "PEAI",
    "ALD-SnO2",
    "PCE",
]
FEATURE_COLUMNS = COLUMNS[3:29]
NUMERIC_COLUMNS = set(
    [
        "NiO2",
        *SOLVENT_COLUMNS,
        "concentration(mg/ml)",
        "wash",
        "E",
        "Cs",
        "FA",
        "MA",
        "Pb",
        "Sn",
        *X_SITE_COLUMNS,
        "C60",
        "BCP",
        "PC60BM",
        "PCBM",
        "PC61BM",
        "PEAI",
        "ALD-SnO2",
        "PCE",
    ]
)

MULTISTEP_PATTERN = re.compile(
    r"分步|逐層|sequential|stepwise|layer\s*[- ]?by\s*[- ]?layer", re.IGNORECASE
)
MULTISTEP_NEGATION_PATTERN = re.compile(
    r"(?:非|不是|並非)\s*(?:分步|逐層)|not\s+(?:sequential|stepwise)", re.IGNORECASE
)

# Legacy workbooks created before the 2026-08-10 sam-dataset-builder update
# may contain values that look valid only because an unrepresented component
# was removed and the visible A/B/X-site fields were renormalized.  The old
# value can sum to one and therefore cannot be detected numerically.  Only
# explicit audit text in Notes/Data_status is used here; no chemical facts are
# guessed from the values themselves.
COMPOSITION_CONTEXT_PATTERN = re.compile(
    r"A[- ]?site|B[- ]?site|X[- ]?site|組成|原式|名義式|前驅鹽|"
    r"Cs/FA|FA/MA|I/Br|PbI|Cs\d|FA\d|MA\d",
    re.IGNORECASE,
)
POSITIVE_NORMALIZATION_PATTERN = re.compile(
    r"(?<!未)(?<!不)(?:重新|重)?正規化|\bnormaliz(?:e|ed|ation)\b",
    re.IGNORECASE,
)
UNREPRESENTED_COMPONENT_PATTERN = re.compile(
    r"排除|剔除|未建模|無對應(?:欄位?|特徵欄)|未設(?:欄位?)|"
    r"含(?:Rb|K|DMA|GA|EDA|PEA)|(?:Rb|K|DMA|GA|EDA|PEA)\s*\d",
    re.IGNORECASE,
)
ORIGINAL_SUM_ANOMALY_PATTERN = re.compile(
    r"(?:原文|原始).{0,100}(?:加總|總和).{0,40}"
    r"(?:\d+(?:\.\d+)?|不等於|偏離|異常|≠)",
    re.IGNORECASE,
)
COMPOSITION_CONFLICT_PATTERN = re.compile(
    r"(?:名義式|名義組成|組成|A[- ]?site|B[- ]?site|X[- ]?site|前驅鹽)"
    r".{0,220}(?:矛盾|衝突|不一致)",
    re.IGNORECASE,
)
COMPOSITION_CONFLICT_NEGATION_PATTERN = re.compile(
    r"(?:無|沒有|未發現|並無).{0,12}(?:矛盾|衝突|不一致)",
    re.IGNORECASE,
)

# Only an explicit two-solvent ratio in Notes is converted automatically.
# Unknown solvents remain in the denominator and therefore normally cause the
# seven represented solvent columns to sum to less than one.
SOLVENT_ALIASES: list[tuple[str, str | None]] = [
    (r"2[- ]?methoxyethanol|2[- ]?me", "2-Methoxyethanol"),
    (r"chlorobenzene|\bcb\b", "chlorobenzene"),
    (r"ch2cl2|dichloromethane|\bdcm\b", "CH2CL2"),
    (r"isopropyl alcohol|isopropanol|\bipa\b", "IPA"),
    (r"tetrahydrofuran|\bthf\b", "THF"),
    (r"anhydrous ethanol|ethanol|\betoh\b|乙醇", "ethanol"),
    (r"toluene|甲苯", "toluene"),
    (r"dimethylformamide|\bdmf\b", None),
    (r"dimethyl sulfoxide|\bdmso\b", None),
    (r"methanol|\bmeoh\b|甲醇", None),
    (r"chloroform|chcl3", None),
    (r"ethyl acetate|乙酸乙酯", None),
    (r"acetone|丙酮", None),
    (r"acetonitrile|乙腈", None),
    (r"dimethylacetamide|\bdmac\b", None),
    (r"\bnmp\b", None),
]
SOLVENT_TOKEN_PATTERN = "(?:" + "|".join(alias for alias, _ in SOLVENT_ALIASES) + ")"
RATIO_PATTERN = re.compile(
    rf"(?P<a>{SOLVENT_TOKEN_PATTERN})\s*[/+:]\s*"
    rf"(?P<b>{SOLVENT_TOKEN_PATTERN})\s*(?:=|為|比例)?\s*"
    r"(?P<x>\d+(?:\.\d+)?)\s*[:/]\s*(?P<y>\d+(?:\.\d+)?)",
    re.IGNORECASE,
)
SOLVENT_CONTEXT_PATTERN = re.compile(
    r"sam|htl|沉積|塗佈|旋塗|浸泡|原液|配製|溶劑", re.IGNORECASE
)

REF_FILENAME_PATTERN = re.compile(
    r"^(?P<prefix>.+)_ref(?P<start>\d+)(?:-(?P<end>\d+))?(?P<suffix>.*)$",
    re.IGNORECASE,
)


class PreprocessError(RuntimeError):
    """Expected user-correctable preprocessing error."""


@dataclass
class SourceRow:
    values: dict[str, Any]
    source: Path
    sheet: str
    excel_row: int
    transformations: list[dict[str, Any]] = field(default_factory=list)

    @property
    def row_id(self) -> str:
        return clean_text(self.values.get("編號"))


@dataclass
class Result:
    kept: list[SourceRow]
    excluded: list[dict[str, Any]]
    transformations: list[dict[str, Any]]
    warnings: list[str]
    input_rows: int


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def decimal_value(value: Any, *, column: str, row: SourceRow) -> Decimal:
    if isinstance(value, bool):
        raise PreprocessError(
            f"{row.source.name} / {row.sheet}!{column}{row.excel_row}: "
            "布林值不能當作數值使用。"
        )
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError, AttributeError) as exc:
        raise PreprocessError(
            f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列 "
            f"欄位 {column!r} 不是合法數值：{value!r}"
        ) from exc
    if not number.is_finite():
        raise PreprocessError(
            f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列 "
            f"欄位 {column!r} 不是有限數值：{value!r}"
        )
    return number


def decimal_string(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(value.quantize(Decimal("1")))
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def display_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return decimal_string(value)
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        return decimal_string(Decimal(str(value)))
    return value


def ensure_inside_project(path: Path, *, label: str) -> Path:
    resolved = path.resolve()
    try:
        resolved.relative_to(PROJECT_ROOT)
    except ValueError as exc:
        raise PreprocessError(
            f"{label}必須位於 {PROJECT_ROOT} 內：{resolved}"
        ) from exc
    return resolved


def resolve_existing_path(argument: str, *, suffix: str, label: str) -> Path:
    supplied = Path(argument).expanduser()
    candidates = (
        [supplied]
        if supplied.is_absolute()
        else [Path.cwd() / supplied, PROJECT_ROOT / supplied, SCRIPT_DIR / supplied]
    )
    matches: list[Path] = []
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            continue
        if resolved.exists() and resolved not in matches:
            matches.append(resolved)
    if not matches:
        raise PreprocessError(f"找不到{label}：{argument}")
    path = matches[0]
    ensure_inside_project(path, label=label)
    if not path.is_file() or path.suffix.lower() != suffix:
        raise PreprocessError(f"{label}必須是 {suffix} 檔案：{path}")
    return path


def resolve_output_path(argument: str | None, inputs: Sequence[Path]) -> Path:
    if argument:
        supplied = Path(argument).expanduser()
        if supplied.is_absolute():
            output = supplied.resolve()
        elif supplied.parent == Path("."):
            output = (CSV_OUTPUT_DIR / supplied.name).resolve()
        else:
            output = (PROJECT_ROOT / supplied).resolve()
    else:
        output = CSV_OUTPUT_DIR / derive_output_name(inputs)
    ensure_inside_project(output, label="輸出檔")
    if output.parent != CSV_OUTPUT_DIR:
        raise PreprocessError(f"輸出 CSV 必須位於 {CSV_OUTPUT_DIR}。")
    if output.suffix.lower() != ".csv":
        raise PreprocessError(f"輸出檔副檔名必須是 .csv：{output}")
    return output


def report_path_for(output: Path) -> Path:
    return JSON_OUTPUT_DIR / f"{output.stem}_preprocess_report.json"


def derive_output_name(inputs: Sequence[Path]) -> str:
    parsed: list[tuple[str, int, int]] = []
    for path in inputs:
        match = REF_FILENAME_PATTERN.match(path.stem)
        if not match:
            raise PreprocessError(
                f"無法從檔名推導輸出名稱：{path.name}；請使用 --output 明確指定。"
            )
        prefix = match.group("prefix")
        start = int(match.group("start"))
        end = int(match.group("end") or start)
        if start > end:
            raise PreprocessError(f"檔名中的 ref 範圍顛倒：{path.name}")
        parsed.append((prefix, start, end))
    prefixes = {item[0] for item in parsed}
    if len(prefixes) != 1:
        raise PreprocessError("多個 XLSX 的 ref 前綴不同；請勿合併不同 review。")
    ordered = sorted(parsed, key=lambda item: (item[1], item[2]))
    for previous, current in zip(ordered, ordered[1:]):
        if current[1] <= previous[2]:
            raise PreprocessError(
                f"XLSX 檔名的 ref 範圍重疊：ref{previous[1]}-{previous[2]} 與 "
                f"ref{current[1]}-{current[2]}"
            )
    prefix = ordered[0][0]
    return f"{prefix}_ref{ordered[0][1]}-{ordered[-1][2]}.csv"


def ref_range(path: Path) -> tuple[int, int]:
    match = REF_FILENAME_PATTERN.match(path.stem)
    if not match:
        return (sys.maxsize, sys.maxsize)
    start = int(match.group("start"))
    return (start, int(match.group("end") or start))


def choose_sheet(workbook: Any, requested: str | None, path: Path) -> str:
    if requested:
        if requested not in workbook.sheetnames:
            raise PreprocessError(
                f"{path.name} 找不到工作表 {requested!r}；可用工作表：{workbook.sheetnames}"
            )
        return requested
    if "主表" in workbook.sheetnames:
        return "主表"
    candidates: list[str] = []
    for worksheet in workbook.worksheets:
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if list(header[: len(COLUMNS)]) == COLUMNS and not any(
            not is_blank(value) for value in header[len(COLUMNS) :]
        ):
            candidates.append(worksheet.title)
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        raise PreprocessError(
            f"{path.name} 沒有 '主表'，也找不到唯一符合固定 35 欄的工作表。"
        )
    raise PreprocessError(
        f"{path.name} 有多個符合 35 欄的工作表 {candidates}；請用 --sheet 指定。"
    )


def read_workbook(path: Path, requested_sheet: str | None) -> tuple[list[SourceRow], list[str]]:
    values_workbook = load_workbook(path, read_only=True, data_only=True)
    formulas_workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_name = choose_sheet(values_workbook, requested_sheet, path)
        if sheet_name not in formulas_workbook.sheetnames:
            raise PreprocessError(f"{path.name} 的公式版本找不到工作表 {sheet_name!r}。")
        values_sheet = values_workbook[sheet_name]
        formulas_sheet = formulas_workbook[sheet_name]
        value_rows = values_sheet.iter_rows(values_only=False)
        formula_rows = formulas_sheet.iter_rows(values_only=False)
        try:
            header_cells = next(value_rows)
            next(formula_rows)
        except StopIteration as exc:
            raise PreprocessError(f"{path.name} / {sheet_name} 是空白工作表。") from exc
        header = [cell.value for cell in header_cells]
        if header[: len(COLUMNS)] != COLUMNS or any(
            not is_blank(value) for value in header[len(COLUMNS) :]
        ):
            actual = [value for value in header if not is_blank(value)]
            raise PreprocessError(
                f"{path.name} / {sheet_name} 欄名或順序不符合固定 35 欄。\n"
                f"預期：{COLUMNS}\n實際：{actual}"
            )

        result: list[SourceRow] = []
        warnings: list[str] = []
        formula_count = 0
        for excel_row, (value_cells, formula_cells) in enumerate(
            zip(value_rows, formula_rows), start=2
        ):
            raw_values = [cell.value for cell in value_cells]
            if any(not is_blank(value) for value in raw_values[len(COLUMNS) :]):
                raise PreprocessError(
                    f"{path.name} / {sheet_name} 第 {excel_row} 列在第 35 欄後仍有資料。"
                )
            values = raw_values[: len(COLUMNS)]
            if not any(not is_blank(value) for value in values):
                continue
            if len(values) < len(COLUMNS):
                values += [None] * (len(COLUMNS) - len(values))
            formula_values = [cell.value for cell in formula_cells[: len(COLUMNS)]]
            for column_index, formula in enumerate(formula_values):
                if isinstance(formula, str) and formula.startswith("="):
                    formula_count += 1
                    if values[column_index] is None:
                        raise PreprocessError(
                            f"{path.name} / {sheet_name} 第 {excel_row} 列欄位 "
                            f"{COLUMNS[column_index]!r} 是公式，但 XLSX 沒有可讀取的快取結果；"
                            "請先用 Excel 重新計算並儲存。"
                        )
            result.append(
                SourceRow(
                    values=dict(zip(COLUMNS, values)),
                    source=path,
                    sheet=sheet_name,
                    excel_row=excel_row,
                )
            )
        if formula_count:
            warnings.append(
                f"{path.name} / {sheet_name} 使用 {formula_count} 個公式的已儲存快取值；"
                "程式本身不重新計算 Excel 公式。"
            )
        return result, warnings
    finally:
        values_workbook.close()
        formulas_workbook.close()


def load_overrides(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None:
        return {}
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            payload = json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        raise PreprocessError(f"無法讀取 overrides JSON：{path}\n{exc}") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("rows"), dict):
        raise PreprocessError("overrides JSON 必須是 {'rows': {'編號': {...}}} 結構。")
    result: dict[str, dict[str, Any]] = {}
    for row_id, instruction in payload["rows"].items():
        if not isinstance(row_id, str) or not row_id.strip() or not isinstance(instruction, dict):
            raise PreprocessError("overrides 的列 ID 與內容格式不合法。")
        reason = instruction.get("reason")
        if not isinstance(reason, str) or not reason.strip():
            raise PreprocessError(f"override {row_id!r} 必須提供非空白 reason。")
        changes = instruction.get("set", {})
        exclude = instruction.get("exclude", False)
        allow_composition_basis = instruction.get("allow_composition_basis", False)
        if changes and exclude:
            raise PreprocessError(f"override {row_id!r} 不可同時 set 與 exclude。")
        if exclude and allow_composition_basis:
            raise PreprocessError(
                f"override {row_id!r} 不可同時 exclude 與 allow_composition_basis。"
            )
        if (
            not isinstance(changes, dict)
            or not isinstance(exclude, bool)
            or not isinstance(allow_composition_basis, bool)
        ):
            raise PreprocessError(
                f"override {row_id!r} 的 set/exclude/allow_composition_basis 格式不合法。"
            )
        for column in changes:
            if column not in COLUMNS or column == "編號":
                raise PreprocessError(
                    f"override {row_id!r} 不允許修改欄位 {column!r}。"
                )
        result[row_id] = {
            "set": changes,
            "exclude": exclude,
            "allow_composition_basis": allow_composition_basis,
            "reason": reason.strip(),
        }
    return result


def alias_to_column(token: str) -> str | None:
    for alias, column in SOLVENT_ALIASES:
        if re.fullmatch(alias, token.strip(), re.IGNORECASE):
            return column
    raise AssertionError(f"Unmatched solvent token: {token}")


def parse_explicit_solvent_ratio(notes: str) -> tuple[dict[str, Decimal], str] | None:
    if not notes or not SOLVENT_CONTEXT_PATTERN.search(notes):
        return None
    match = RATIO_PATTERN.search(notes)
    if not match:
        return None
    first = Decimal(match.group("x"))
    second = Decimal(match.group("y"))
    denominator = first + second
    if denominator <= 0:
        return None
    values = {column: Decimal("0") for column in SOLVENT_COLUMNS}
    first_column = alias_to_column(match.group("a"))
    second_column = alias_to_column(match.group("b"))
    if first_column:
        values[first_column] += first / denominator
    if second_column:
        values[second_column] += second / denominator
    return values, match.group(0)


def composition_basis_issue(row: SourceRow) -> dict[str, str] | None:
    """Return an explicit legacy composition warning recorded in audit text.

    A sum of one is not evidence that the original basis was complete.  This
    detector therefore relies only on Notes/Data_status phrases that explicitly
    disclose renormalization after omitted components, an anomalous original
    sum, or a source-composition conflict.
    """

    audit_text = "；".join(
        value
        for value in (
            clean_text(row.values.get("Notes")),
            clean_text(row.values.get("Data_status")),
        )
        if value
    )
    if not audit_text:
        return None

    conflict = COMPOSITION_CONFLICT_PATTERN.search(audit_text)
    if conflict and not COMPOSITION_CONFLICT_NEGATION_PATTERN.search(audit_text):
        return {
            "signal": "composition_source_conflict",
            "matched_text": conflict.group(0),
            "audit_text": audit_text[:600],
        }

    if not COMPOSITION_CONTEXT_PATTERN.search(audit_text):
        return None
    normalization = POSITIVE_NORMALIZATION_PATTERN.search(audit_text)
    if not normalization:
        return None
    if UNREPRESENTED_COMPONENT_PATTERN.search(audit_text):
        return {
            "signal": "renormalized_after_unrepresented_component",
            "matched_text": normalization.group(0),
            "audit_text": audit_text[:600],
        }
    if ORIGINAL_SUM_ANOMALY_PATTERN.search(audit_text):
        return {
            "signal": "renormalized_original_sum_anomaly",
            "matched_text": normalization.group(0),
            "audit_text": audit_text[:600],
        }
    return None


def exclusion(row: SourceRow, code: str, detail: Any) -> dict[str, Any]:
    return {
        "id": row.row_id,
        "source": row.source.name,
        "sheet": row.sheet,
        "excel_row": row.excel_row,
        "code": code,
        "detail": detail,
    }


def transformation(row: SourceRow, code: str, detail: Any) -> dict[str, Any]:
    item = {
        "id": row.row_id,
        "source": row.source.name,
        "sheet": row.sheet,
        "excel_row": row.excel_row,
        "code": code,
        "detail": detail,
    }
    row.transformations.append(item)
    return item


def validate_identity(rows: Sequence[SourceRow]) -> None:
    seen: dict[str, SourceRow] = {}
    for row in rows:
        if not row.row_id:
            raise PreprocessError(
                f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列缺少 '編號'。"
            )
        if row.row_id in seen:
            previous = seen[row.row_id]
            raise PreprocessError(
                f"重複編號 {row.row_id!r}：{previous.source.name} 第 {previous.excel_row} 列，"
                f"以及 {row.source.name} 第 {row.excel_row} 列。"
            )
        seen[row.row_id] = row


def apply_overrides(
    rows: Sequence[SourceRow], overrides: dict[str, dict[str, Any]]
) -> tuple[set[str], set[str], list[dict[str, Any]]]:
    if not overrides:
        return set(), set(), []
    by_id = {row.row_id: row for row in rows}
    unknown = sorted(set(overrides) - set(by_id))
    if unknown:
        raise PreprocessError(f"overrides 含有找不到的編號：{unknown}")
    manual_exclusions: set[str] = set()
    composition_basis_approvals: set[str] = set()
    transformations: list[dict[str, Any]] = []
    for row_id, instruction in overrides.items():
        row = by_id[row_id]
        if instruction["exclude"]:
            manual_exclusions.add(row_id)
            continue
        if instruction["allow_composition_basis"]:
            composition_basis_approvals.add(row_id)
            transformations.append(
                transformation(
                    row,
                    "manual_composition_basis_approval",
                    {"reason": instruction["reason"]},
                )
            )
        before = {column: display_value(row.values.get(column)) for column in instruction["set"]}
        row.values.update(instruction["set"])
        after = {column: display_value(row.values.get(column)) for column in instruction["set"]}
        transformations.append(
            transformation(
                row,
                "manual_override",
                {"reason": instruction["reason"], "before": before, "after": after},
            )
        )
    return manual_exclusions, composition_basis_approvals, transformations


def validate_numeric_contract(row: SourceRow) -> None:
    for column in NUMERIC_COLUMNS:
        value = row.values.get(column)
        if is_blank(value):
            if column in REQUIRED_COLUMNS:
                continue
            continue
        row.values[column] = decimal_value(value, column=column, row=row)

    for column in BINARY_COLUMNS:
        value = row.values.get(column)
        if is_blank(value):
            continue
        if value not in (Decimal("0"), Decimal("1")):
            raise PreprocessError(
                f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列 "
                f"{column} 必須是 0 或 1，目前為 {display_value(value)!r}。"
            )
    for column in SOLVENT_COLUMNS:
        value = row.values[column]
        if value < 0 or value > 1:
            raise PreprocessError(
                f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列 "
                f"{column} 必須介於 0 與 1，目前為 {display_value(value)!r}。"
            )
    pce = row.values["PCE"]
    if pce <= 0 or pce > 35:
        raise PreprocessError(
            f"{row.source.name} / {row.sheet} 第 {row.excel_row} 列 PCE 必須滿足 "
            f"0 < PCE <= 35，目前為 {display_value(pce)!r}。"
        )


def preprocess(rows: list[SourceRow], overrides: dict[str, dict[str, Any]]) -> Result:
    validate_identity(rows)
    (
        manual_exclusions,
        composition_basis_approvals,
        transformations,
    ) = apply_overrides(rows, overrides)
    excluded: list[dict[str, Any]] = []
    warnings: list[str] = []
    candidates: list[SourceRow] = []

    # Step 0: legacy audit text can reveal a hidden composition-basis problem
    # even when the visible numbers sum to one.  This must precede the ordinary
    # missing-value checks so the report preserves the scientifically relevant
    # exclusion reason (for example, a row may also have blank solvent fields).
    for row in rows:
        if row.row_id in manual_exclusions:
            excluded.append(
                exclusion(row, "manual_exclusion", overrides[row.row_id]["reason"])
            )
            continue
        basis_issue = (
            None
            if row.row_id in composition_basis_approvals
            else composition_basis_issue(row)
        )
        if basis_issue:
            excluded.append(
                exclusion(row, "composition_basis_manual_review", basis_issue)
            )
            continue

        # Step 1: required fields. Numeric 0 is not blank.
        missing = [column for column in REQUIRED_COLUMNS if is_blank(row.values.get(column))]
        if missing:
            excluded.append(exclusion(row, "missing_required", {"columns": missing}))
            continue
        validate_numeric_contract(row)
        candidates.append(row)

    # Step 2: deterministic solvent conversion/filtering.
    solvent_candidates: list[SourceRow] = []
    for row in candidates:
        notes = clean_text(row.values.get("Notes"))
        ratio = parse_explicit_solvent_ratio(notes)
        if ratio:
            converted, matched_text = ratio
            before = {column: display_value(row.values[column]) for column in SOLVENT_COLUMNS}
            after = {column: display_value(value) for column, value in converted.items()}
            if before != after:
                row.values.update(converted)
                transformations.append(
                    transformation(
                        row,
                        "explicit_solvent_ratio",
                        {"matched_text": matched_text, "before": before, "after": after},
                    )
                )
        solvent_total = sum(row.values[column] for column in SOLVENT_COLUMNS)
        text = f"{row.row_id} {notes}"
        if MULTISTEP_PATTERN.search(text) and not MULTISTEP_NEGATION_PATTERN.search(text):
            excluded.append(
                exclusion(
                    row,
                    "multistep_solvent_process",
                    {"solvent_sum": decimal_string(solvent_total)},
                )
            )
            continue
        if abs(solvent_total - Decimal("1")) > TOLERANCE:
            excluded.append(
                exclusion(
                    row,
                    "solvent_sum_not_one",
                    {
                        "solvent_sum": decimal_string(solvent_total),
                        "values": {
                            column: display_value(row.values[column])
                            for column in SOLVENT_COLUMNS
                        },
                    },
                )
            )
            continue
        solvent_candidates.append(row)

    # Step 3: ETL all-zero exclusion. A/B non-unit sums are retained but reported.
    etl_candidates: list[SourceRow] = []
    for row in solvent_candidates:
        etl_total = sum(row.values[column] for column in ETL_COLUMNS)
        if abs(etl_total) <= TOLERANCE:
            excluded.append(
                exclusion(
                    row,
                    "etl_all_zero",
                    {"values": {column: display_value(row.values[column]) for column in ETL_COLUMNS}},
                )
            )
            continue
        for label, columns in (("A_site", A_SITE_COLUMNS), ("B_site", B_SITE_COLUMNS)):
            total = sum(row.values[column] for column in columns)
            if abs(total - Decimal("1")) > TOLERANCE:
                warnings.append(
                    f"{row.row_id} {label}總和={decimal_string(total)}；依規則保留，不自動刪除。"
                )
        etl_candidates.append(row)

    # Step 4: X-site basis conversion and anomaly filtering.
    x_candidates: list[SourceRow] = []
    for row in etl_candidates:
        x_total = sum(row.values[column] for column in X_SITE_COLUMNS)
        if abs(x_total - Decimal("3")) <= TOLERANCE:
            before = {column: display_value(row.values[column]) for column in X_SITE_COLUMNS}
            for column in X_SITE_COLUMNS:
                row.values[column] = row.values[column] / Decimal("3")
            after = {column: display_value(row.values[column]) for column in X_SITE_COLUMNS}
            transformations.append(
                transformation(
                    row,
                    "x_site_divide_by_three",
                    {"before": before, "after": after},
                )
            )
            x_total = sum(row.values[column] for column in X_SITE_COLUMNS)
        if abs(x_total - Decimal("1")) > TOLERANCE:
            excluded.append(
                exclusion(
                    row,
                    "x_site_sum_anomaly_manual_review",
                    {
                        "x_site_sum": decimal_string(x_total),
                        "values": {
                            column: display_value(row.values[column])
                            for column in X_SITE_COLUMNS
                        },
                    },
                )
            )
            continue
        x_candidates.append(row)

    # Step 5: SMILES + 26 features with different PCE -> keep the highest PCE.
    groups: dict[tuple[str, ...], list[SourceRow]] = defaultdict(list)
    for row in x_candidates:
        key = tuple(
            clean_text(display_value(row.values.get(column)))
            for column in ["smile", *FEATURE_COLUMNS]
        )
        groups[key].append(row)

    deduplicated: list[SourceRow] = []
    for group in groups.values():
        if len(group) == 1:
            deduplicated.extend(group)
            continue
        pce_values = {row.values["PCE"] for row in group}
        if len(pce_values) == 1:
            warnings.append(
                "相同 SMILES + 26 特徵且 PCE 相同，依規則未自動刪除："
                + ", ".join(row.row_id for row in group)
            )
            deduplicated.extend(group)
            continue
        maximum = max(pce_values)
        winners = [row for row in group if row.values["PCE"] == maximum]
        if len(winners) != 1:
            raise PreprocessError(
                "相同 SMILES + 26 特徵有多筆並列最高 PCE，無法無歧義擇一："
                + ", ".join(row.row_id for row in winners)
            )
        winner = winners[0]
        deduplicated.append(winner)
        for row in group:
            if row is winner:
                continue
            excluded.append(
                exclusion(
                    row,
                    "duplicate_features_lower_pce",
                    {
                        "kept_id": winner.row_id,
                        "kept_pce": display_value(winner.values["PCE"]),
                        "excluded_pce": display_value(row.values["PCE"]),
                    },
                )
            )

    kept_ids = {id(row) for row in deduplicated}
    kept = [row for row in rows if id(row) in kept_ids]
    return Result(
        kept=kept,
        excluded=excluded,
        transformations=transformations,
        warnings=warnings,
        input_rows=len(rows),
    )


def count_by_code(items: Iterable[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for item in items:
        counts[item["code"]] += 1
    return dict(sorted(counts.items()))


def build_report(
    result: Result,
    inputs: Sequence[Path],
    output: Path,
    report_path: Path,
    overrides_path: Path | None,
    workbook_warnings: Sequence[str],
) -> dict[str, Any]:
    warnings = [*workbook_warnings, *result.warnings]
    return {
        "script": "first_preprocessed_csv/first_preprocess.py",
        "script_version": VERSION,
        "rules": "first_preprocessed_csv/FIRST_PREPROCESSING_GUIDE.md",
        "tolerance": decimal_string(TOLERANCE),
        "inputs": [str(path.relative_to(PROJECT_ROOT)) for path in inputs],
        "output": str(output.relative_to(PROJECT_ROOT)),
        "report": str(report_path.relative_to(PROJECT_ROOT)),
        "overrides": (
            str(overrides_path.relative_to(PROJECT_ROOT)) if overrides_path else None
        ),
        "summary": {
            "input_rows": result.input_rows,
            "kept_rows": len(result.kept),
            "excluded_rows": len(result.excluded),
            "excluded_by_code": count_by_code(result.excluded),
            "transformations": len(result.transformations),
            "transformations_by_code": count_by_code(result.transformations),
            "warnings": len(warnings),
        },
        "excluded": result.excluded,
        "transformations": result.transformations,
        "warnings": warnings,
    }


def atomic_write_csv(path: Path, rows: Sequence[SourceRow], overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise PreprocessError(f"輸出檔已存在；如要取代請加 --overwrite：{path}")
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8-sig",
            newline="",
            dir=path.parent,
            prefix=f".{path.stem}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
            writer = csv.writer(handle, quoting=csv.QUOTE_ALL, lineterminator="\n")
            writer.writerow(COLUMNS)
            for row in rows:
                writer.writerow([display_value(row.values.get(column)) for column in COLUMNS])
        os.replace(temporary, path)
    finally:
        if temporary and temporary.exists():
            temporary.unlink()


def atomic_write_json(path: Path, payload: dict[str, Any], overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise PreprocessError(f"報告檔已存在；如要取代請加 --overwrite：{path}")
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=path.parent,
            prefix=f".{path.stem}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        os.replace(temporary, path)
    finally:
        if temporary and temporary.exists():
            temporary.unlink()


def summary_lines(report: dict[str, Any]) -> list[str]:
    summary = report["summary"]
    lines = [
        f"輸入：{summary['input_rows']} 列",
        f"保留：{summary['kept_rows']} 列",
        f"排除：{summary['excluded_rows']} 列",
        f"轉換：{summary['transformations']} 項",
    ]
    if summary["excluded_by_code"]:
        lines.append(
            "排除原因："
            + "、".join(
                f"{code}={count}"
                for code, count in summary["excluded_by_code"].items()
            )
        )
    if summary["warnings"]:
        lines.append(f"警告：{summary['warnings']} 項（詳見 JSON 報告）")
    return lines


def discover_workbook_folders() -> list[tuple[Path, list[Path]]]:
    folders: list[tuple[Path, list[Path]]] = []
    for folder in sorted(PROJECT_ROOT.iterdir(), key=lambda path: path.name.casefold()):
        if not folder.is_dir() or not PAPER_DIR_PATTERN.fullmatch(folder.name):
            continue
        workbooks = sorted(
            (
                path
                for path in folder.glob("*.xlsx")
                if path.is_file() and not path.name.startswith("~$")
            ),
            key=lambda path: path.name.casefold(),
        )
        if workbooks:
            folders.append((folder, workbooks))
    return folders


def parse_number_selection(raw: str, item_count: int) -> list[int]:
    text = raw.strip()
    if not text:
        raise PreprocessError("請至少輸入一個編號。")
    selected: set[int] = set()
    for token in re.split(r"[,，\s]+", text):
        if not token:
            continue
        match = re.fullmatch(r"(\d+)(?:\s*-\s*(\d+))?", token)
        if not match:
            raise PreprocessError(f"無法辨識的選擇：{token!r}")
        start = int(match.group(1))
        end = int(match.group(2) or start)
        if start > end:
            raise PreprocessError(f"編號範圍顛倒：{token}")
        if start < 1 or end > item_count:
            raise PreprocessError(f"編號必須介於 1 到 {item_count}：{token}")
        selected.update(range(start, end + 1))
    return sorted(selected)


def prompt_yes_no(message: str) -> bool:
    while True:
        answer = input(f"{message} [y/N]：").strip().casefold()
        if answer in {"y", "yes", "是"}:
            return True
        if answer in {"", "n", "no", "否"}:
            return False
        print("請輸入 y 或 n。")


def interactive_main() -> int:
    print("第一次前處理互動模式")
    print("輸入 q 可隨時取消；來源 XLSX 不會被修改。\n")
    folders = discover_workbook_folders()
    if not folders:
        raise PreprocessError("找不到符合 YYYY_* 命名且直接包含 XLSX 的論文資料夾。")

    print("請選擇論文資料夾：")
    for index, (folder, workbooks) in enumerate(folders, start=1):
        print(f"  {index}. {folder.name}（{len(workbooks)} 個 XLSX）")
    raw_folder = input("資料夾編號：").strip()
    if raw_folder.casefold() in {"q", "quit", "exit"}:
        print("已取消，未寫入任何檔案。")
        return 0
    folder_indexes = parse_number_selection(raw_folder, len(folders))
    if len(folder_indexes) != 1:
        raise PreprocessError("論文資料夾一次只能選擇一個。")
    folder, workbooks = folders[folder_indexes[0] - 1]

    print(f"\n{folder.name} 內可用的 XLSX：")
    for index, workbook in enumerate(workbooks, start=1):
        print(f"  {index}. {workbook.name}")
    raw_workbooks = input("XLSX 編號（可輸入 1,3-5）：").strip()
    if raw_workbooks.casefold() in {"q", "quit", "exit"}:
        print("已取消，未寫入任何檔案。")
        return 0
    workbook_indexes = parse_number_selection(raw_workbooks, len(workbooks))
    selected = [workbooks[index - 1] for index in workbook_indexes]

    print("\n已選擇：")
    for workbook in selected:
        print(f"  - {workbook.relative_to(PROJECT_ROOT)}")
    selected_arguments = [str(path) for path in selected]

    print("\n開始試跑（不寫入檔案）……")
    dry_run_result = main([*selected_arguments, "--dry-run"])
    if dry_run_result != 0:
        print("試跑未通過，未寫入任何檔案。")
        return dry_run_result
    if not prompt_yes_no("試跑已通過，是否正式產出 CSV 與 JSON？"):
        print("已取消正式寫入。")
        return 0

    output = resolve_output_path(None, selected)
    report_path = report_path_for(output)
    existing = [path for path in (output, report_path) if path.exists()]
    arguments = list(selected_arguments)
    if existing:
        print("\n下列輸出已存在：")
        for path in existing:
            print(f"  - {path.relative_to(PROJECT_ROOT)}")
        if not prompt_yes_no("是否原子式覆寫上述既有輸出？"):
            print("已取消覆寫，既有檔案未變更。")
            return 0
        arguments.append("--overwrite")

    print("\n開始正式寫入……")
    return main(arguments)


def parse_arguments(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "依 FIRST_PREPROCESSING_GUIDE.md 對明確指定的 XLSX 執行第一次前處理，"
            "並將 CSV 與稽核 JSON 分別寫入 first_preprocessed_csv/csv 與 "
            "first_preprocessed_csv/json。"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "範例：\n"
            "  .venv\\Scripts\\python.exe first_preprocessed_csv\\first_preprocess.py "
            '"2025_Suo_Review\\2025_Suo_Review_ref1-101_GPT-5.6-sol.xlsx"\n'
            "  .venv\\Scripts\\python.exe first_preprocessed_csv\\first_preprocess.py part1.xlsx part2.xlsx "
            "--output combined_ref1-100.csv\n"
            "\n"
            "overrides JSON：\n"
            "  {\n"
            '    "rows": {\n'
            '      "63-control(CbzNaph)": {\n'
            '        "set": {"ethanol": 1},\n'
            '        "reason": "SI Methods 明示 ethanol"\n'
            "      }\n"
            "    }\n"
            "  }"
        ),
    )
    parser.add_argument(
        "xlsx",
        nargs="*",
        help="一個或多個明確指定的 XLSX；不接受只輸入資料夾。",
    )
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="以兩層編號選單選擇論文資料夾及一個或多個 XLSX。",
    )
    parser.add_argument("--output", help="輸出 CSV 檔名；省略時由 ref 範圍推導。")
    parser.add_argument("--sheet", help="主資料工作表名稱；預設優先使用 '主表'。")
    parser.add_argument(
        "--overrides",
        help="可選的人工裁決 JSON；每項 set/exclude 都必須附 reason。",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="允許原子式取代既有 CSV 與報告。",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="完成讀取、轉換與驗證，但不寫入任何檔案。",
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    try:
        args = parse_arguments(argv)
        if args.interactive:
            if (
                args.xlsx
                or args.output
                or args.sheet
                or args.overrides
                or args.overwrite
                or args.dry_run
            ):
                raise PreprocessError("--interactive 不可與 XLSX 或其他處理參數併用。")
            return interactive_main()
        if not args.xlsx:
            raise PreprocessError("請明確指定一個或多個 XLSX，或使用 --interactive。")
        inputs = [
            resolve_existing_path(item, suffix=".xlsx", label="輸入 XLSX")
            for item in args.xlsx
        ]
        if len(set(inputs)) != len(inputs):
            raise PreprocessError("同一個 XLSX 被重複指定。")
        input_dirs = {path.parent for path in inputs}
        if len(input_dirs) != 1:
            raise PreprocessError("所有輸入 XLSX 必須位於同一個內層資料夾。")
        output = resolve_output_path(args.output, inputs)
        report_path = report_path_for(output)
        if output in inputs:
            raise PreprocessError("輸出 CSV 不可與輸入 XLSX 相同。")

        overrides_path = (
            resolve_existing_path(args.overrides, suffix=".json", label="overrides JSON")
            if args.overrides
            else None
        )
        overrides = load_overrides(overrides_path)

        sorted_inputs = sorted(inputs, key=ref_range)
        rows: list[SourceRow] = []
        workbook_warnings: list[str] = []
        for path in sorted_inputs:
            workbook_rows, warnings = read_workbook(path, args.sheet)
            rows.extend(workbook_rows)
            workbook_warnings.extend(warnings)

        result = preprocess(rows, overrides)
        report = build_report(
            result, sorted_inputs, output, report_path, overrides_path, workbook_warnings
        )
        for line in summary_lines(report):
            print(line)

        if args.dry_run:
            print("Dry run：未寫入 CSV 或 JSON 報告。")
            return 0

        if report_path.exists() and not args.overwrite:
            raise PreprocessError(
                f"報告檔已存在；如要取代請加 --overwrite：{report_path}"
            )
        CSV_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        JSON_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        atomic_write_csv(output, result.kept, args.overwrite)
        atomic_write_json(report_path, report, args.overwrite)
        print(f"CSV：{output}")
        print(f"稽核報告：{report_path}")
        return 0
    except PreprocessError as exc:
        print(f"錯誤：{exc}", file=sys.stderr)
        return 2
    except KeyboardInterrupt:
        print("已中止；未完成的暫存輸出不會取代正式檔案。", file=sys.stderr)
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
